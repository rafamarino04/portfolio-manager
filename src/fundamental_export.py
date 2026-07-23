"""
Export in Excel dell'Analisi Fondamentale v2.0 (src/fundamental_score.py):
gli stessi elementi mostrati in pagina — i due assi Quality/Valuation
separati, la matrice 2x2, l'archetipo/Dickinson, Piotroski/Altman/
Beneish, le Note Critiche selettive, il modello di confidenza — ma
scaricabili e riutilizzabili fuori dall'app. Nessun foglio "peer group":
il motore v2.0 non costruisce più un peer group a runtime (scoring
assoluto per settore/archetipo, non percentile).
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import financials as finmod
from src import fundamental_score as fscore

NAVY = "1B2A4A"
GOLD = "C9A227"
GREEN = "1E8E5A"
RED = "C0392B"
GRAY = "6B7280"
INPUT_BLUE = "0000FF"
HEADER_FILL = "1B2A4A"

FONT_NAME = "Arial"

_RAW_ROWS = [
    ("revenue", "Ricavi"), ("ebitda", "EBITDA"), ("gross_profit", "Utile lordo"),
    ("operating_income", "Utile operativo (EBIT)"), ("net_income", "Utile netto"),
    ("free_cash_flow", "Free cash flow"), ("total_debt", "Debito totale"),
    ("cash", "Cassa"), ("total_equity", "Patrimonio netto"), ("total_assets", "Attivo totale"),
    ("retained_earnings", "Utili non distribuiti"), ("eps", "EPS"),
]

_METRIC_LABELS_IT = {
    "roic": "ROIC %", "gross_profits_to_assets": "Gross profit / Attivo %",
    "operating_margin_current": "Margine operativo %", "shareholder_yield": "Shareholder yield %",
    "fcf_conversion": "FCF conversion %", "accruals_ratio": "Accruals ratio (Sloan) %",
    "net_debt_to_ebitda": "Debito netto / EBITDA (x)", "interest_coverage": "Copertura interessi (x)",
    "revenue_cagr": "CAGR ricavi %", "eps_cagr": "CAGR EPS %", "growth_volatility": "Volatilità crescita ricavi %",
}
_METRIC_CATEGORY = {
    "roic": "profitability", "gross_profits_to_assets": "profitability",
    "operating_margin_current": "profitability", "shareholder_yield": "profitability",
    "fcf_conversion": "earnings_quality", "accruals_ratio": "earnings_quality",
    "net_debt_to_ebitda": "financial_strength", "interest_coverage": "financial_strength",
    "revenue_cagr": "growth_quality", "eps_cagr": "growth_quality", "growth_volatility": "growth_quality",
}


def _header_font():
    return Font(name=FONT_NAME, bold=True, color="FFFFFF")


def _title_font(size=14):
    return Font(name=FONT_NAME, bold=True, size=size, color=NAVY)


def _label_font(bold=False):
    return Font(name=FONT_NAME, bold=bold, color=NAVY)


def _input_font():
    return Font(name=FONT_NAME, color=INPUT_BLUE)


def _formula_font():
    return Font(name=FONT_NAME, color="000000")


def _style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = _header_font()
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _band_color(band: str) -> str:
    return {"Eccellente": GREEN, "Buono": GREEN, "Discreto": GOLD, "Sufficiente": GOLD,
            "Debole": RED, "Scarso": RED}.get(band, GRAY)


# ---------------------------------------------------------------------------
# Foglio 1 — Sintesi
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws, symbol: str, info: dict, price, result: dict):
    ws.cell(row=1, column=1, value=f"Analisi Fondamentale v2.0 — {info.get('name', symbol)} ({symbol})").font = _title_font(16)
    ws.cell(row=2, column=1, value=(
        "Portfolio Manager · dati Yahoo Finance (yfinance) · scoring assoluto per settore/archetipo, "
        "nessun peer group a runtime · solo a scopo informativo, non consulenza finanziaria"
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    row = 4
    facts = [
        ("Prezzo", price, "$#,##0.00;($#,##0.00);-"),
        ("Settore", info.get("sector"), None),
        ("Bucket di soglie", result.get("bucket_label"), None),
        ("Archetipo operativo", result.get("archetype_label"), None),
        ("Stadio Dickinson", result.get("dickinson_latest_label"), None),
        ("Cap bucket", result.get("cap_bucket"), None),
        ("Capitalizzazione", info.get("market_cap"), "$#,##0,,\"M\""),
        ("Valuta", info.get("currency"), None),
        ("Affidabilità analisi", result.get("confidence", {}).get("level"), None),
    ]
    for label, val, fmt in facts:
        ws.cell(row=row, column=1, value=label).font = _label_font(bold=True)
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = _input_font()
        if fmt:
            cell.number_format = fmt
        row += 1

    row += 1
    quality, valuation = result["quality"], result["valuation"]
    ws.cell(row=row, column=1, value="Quality (asse 1)").font = _title_font(12)
    row += 1
    q_score = quality.get("score")
    ws.cell(row=row, column=1, value="Punteggio (0-100)").font = _label_font()
    cell = ws.cell(row=row, column=2, value=q_score)
    cell.font = Font(name=FONT_NAME, bold=True, color=_band_color(quality.get("band", "n/d")))
    cell.number_format = "0"
    row += 1
    ws.cell(row=row, column=1, value="Banda").font = _label_font()
    ws.cell(row=row, column=2, value=quality.get("band", "n/d")).font = Font(
        name=FONT_NAME, bold=True, color=_band_color(quality.get("band", "n/d")))
    row += 2

    ws.cell(row=row, column=1, value="Valuation (asse 2, alto = economico)").font = _title_font(12)
    row += 1
    v_score = valuation.get("score")
    ws.cell(row=row, column=1, value="Punteggio (0-100)").font = _label_font()
    cell = ws.cell(row=row, column=2, value=v_score)
    cell.font = Font(name=FONT_NAME, bold=True, color=_band_color(valuation.get("band", "n/d")))
    cell.number_format = "0"
    row += 1
    ws.cell(row=row, column=1, value="Banda").font = _label_font()
    ws.cell(row=row, column=2, value=valuation.get("band", "n/d")).font = Font(
        name=FONT_NAME, bold=True, color=_band_color(valuation.get("band", "n/d")))
    row += 2

    matrix = result.get("matrix")
    if matrix:
        ws.cell(row=row, column=1, value="Quadrante matrice Quality x Valuation").font = _title_font(11)
        row += 1
        ws.cell(row=row, column=1, value=matrix["label"]).font = Font(name=FONT_NAME, bold=True, color=NAVY)
        row += 1
        ws.cell(row=row, column=1, value=matrix["action"]).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)
        row += 2

    if result.get("blended") is not None:
        ws.cell(row=row, column=1, value="Numero unico secondario (media Quality/Valuation)").font = _label_font()
        c = ws.cell(row=row, column=2, value=result["blended"])
        c.number_format = "0"
        c.font = _input_font()
        row += 2

    ws.cell(row=row, column=1, value="Badge").font = _title_font(11)
    row += 1
    piotroski = result["piotroski"]
    ws.cell(row=row, column=1, value="Piotroski F-Score (0-9)").font = _label_font()
    ws.cell(row=row, column=2, value=piotroski.get("score")).font = _input_font()
    row += 1
    altman = result["altman"]
    ws.cell(row=row, column=1, value=f"Altman {altman.get('variant') or 'Z'}").font = _label_font()
    a_cell = ws.cell(row=row, column=2, value=altman.get("z"))
    a_cell.font = _input_font()
    a_cell.number_format = "0.00"
    row += 1
    zone_it = {"safe": "Sicura", "grey": "Grigia", "distress": "Distress"}.get(altman.get("zone"), "n/d")
    ws.cell(row=row, column=1, value="Zona Altman").font = _label_font()
    z_cell = ws.cell(row=row, column=2, value=zone_it)
    z_cell.font = Font(name=FONT_NAME, bold=True, color={"Sicura": GREEN, "Grigia": GOLD, "Distress": RED}.get(zone_it, GRAY))
    row += 1
    beneish = result.get("beneish", {})
    ws.cell(row=row, column=1, value=f"Beneish M-Score ({beneish.get('version') or 'n/d'})").font = _label_font()
    m_cell = ws.cell(row=row, column=2, value=beneish.get("m_score"))
    m_cell.font = _input_font()
    m_cell.number_format = "0.00"
    row += 2

    ws.cell(row=row, column=1, value="Tesi in una riga").font = _title_font(11)
    row += 1
    ws.cell(row=row, column=1, value=fscore.build_thesis_text(result))
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=6)
    ws.row_dimensions[row].height = 50
    row += 4

    bulls, bears = fscore.build_bull_bear(result)
    ws.cell(row=row, column=1, value="Punti di forza").font = Font(name=FONT_NAME, bold=True, color=GREEN)
    row += 1
    if bulls:
        for b in bulls:
            ws.cell(row=row, column=1, value=f"+ {b}")
            row += 1
    else:
        ws.cell(row=row, column=1, value="Nessun punto di forza netto in assoluto.").font = Font(
            name=FONT_NAME, italic=True, size=9, color=GRAY)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Punti di attenzione").font = Font(name=FONT_NAME, bold=True, color=RED)
    row += 1
    if bears:
        for b in bears:
            ws.cell(row=row, column=1, value=f"- {b}")
            row += 1
    else:
        ws.cell(row=row, column=1, value="Nessun segnale di attenzione rilevato.").font = Font(
            name=FONT_NAME, italic=True, size=9, color=GRAY)
        row += 1

    _autosize(ws, {"A": 46, "B": 18, "C": 14, "D": 14, "E": 14, "F": 14})


# ---------------------------------------------------------------------------
# Foglio 2 — Metriche core (per categoria, punteggio assoluto)
# ---------------------------------------------------------------------------

def _write_metrics_sheet(ws, symbol: str, result: dict):
    ws.cell(row=1, column=1, value=f"Metriche core — {symbol}").font = _title_font(14)
    ws.cell(row=2, column=1, value=(
        "Valore grezzo per metrica, raggruppato per categoria dell'asse Quality — punteggio assoluto "
        "su scala fissa (src/sector_thresholds.py), non un percentile contro altri titoli."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    header_row = 4
    for j, h in enumerate(["Metrica", "Categoria", "Valore"]):
        ws.cell(row=header_row, column=1 + j, value=h)
    _style_header_row(ws, header_row, 3)

    metrics = result.get("metrics", {})
    row = header_row + 1
    for key, label in _METRIC_LABELS_IT.items():
        ws.cell(row=row, column=1, value=label).font = _label_font()
        ws.cell(row=row, column=2, value=fscore.CATEGORY_LABELS_IT.get(_METRIC_CATEGORY.get(key), "n/d")).font = _label_font()
        val_cell = ws.cell(row=row, column=3, value=metrics.get(key))
        val_cell.font = _input_font()
        val_cell.number_format = "0.00"
        row += 1

    ws.cell(row=row + 1, column=1, value=(
        "Fonte: src/fundamental_score.py sui bilanci storici Yahoo Finance. Le soglie assolute per "
        "settore/archetipo sono in src/sector_thresholds.py."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    _autosize(ws, {"A": 34, "B": 30, "C": 14})


# ---------------------------------------------------------------------------
# Foglio 3 — Categorie e pesi (Quality, ricalcolabile)
# ---------------------------------------------------------------------------

def _write_categories_sheet(ws, symbol: str, result: dict):
    ws.cell(row=1, column=1, value=f"Categorie, pesi e punteggio Quality — {symbol}").font = _title_font(14)
    ws.cell(row=2, column=1, value=(
        f"Pesi determinati dall'archetipo operativo ({result.get('archetype_label')}), non dal solo "
        "settore GICS (bug-fix v2.0)."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    header_row = 4
    for j, h in enumerate(["Categoria", "Punteggio assoluto (0-100)", "Peso nel composito (%)"]):
        ws.cell(row=header_row, column=1 + j, value=h)
    _style_header_row(ws, header_row, 3)

    quality = result.get("quality", {})
    subscores = quality.get("subscores", {})
    weights_used = quality.get("category_weights_used", {})
    row = header_row + 1
    first_cat_row = row
    for cat in fscore.CATEGORIES:
        ws.cell(row=row, column=1, value=fscore.CATEGORY_LABELS_IT[cat]).font = _label_font()
        s_cell = ws.cell(row=row, column=2, value=subscores.get(cat))
        s_cell.font = _input_font()
        s_cell.number_format = "0.0"
        w_cell = ws.cell(row=row, column=3, value=weights_used.get(cat))
        w_cell.font = _input_font()
        w_cell.number_format = "0.0"
        row += 1
    last_cat_row = row - 1

    piotroski = result.get("piotroski", {})
    piotroski_weight = quality.get("piotroski_weight_used")
    ws.cell(row=row, column=1, value="Piotroski F-Score (scalato 0-100)").font = _label_font()
    piotroski_scaled = (piotroski["score"] / 9 * 100) if piotroski.get("score") is not None else None
    p_cell = ws.cell(row=row, column=2, value=piotroski_scaled)
    p_cell.font = _input_font()
    p_cell.number_format = "0.0"
    pw_cell = ws.cell(row=row, column=3, value=piotroski_weight)
    pw_cell.font = _input_font()
    pw_cell.number_format = "0.0"
    piotroski_row = row
    row += 2

    ws.cell(row=row, column=1, value="Punteggio Quality (formula ricalcolabile)").font = _label_font(bold=True)
    total_cell = ws.cell(
        row=row, column=2,
        value=f"=IFERROR(SUMPRODUCT(B{first_cat_row}:B{last_cat_row},C{first_cat_row}:C{last_cat_row})/100"
              f"+IF(B{piotroski_row}=\"\",0,B{piotroski_row}*C{piotroski_row}/100),\"n/d\")",
    )
    total_cell.font = Font(name=FONT_NAME, bold=True, color=NAVY)
    total_cell.number_format = "0.0"
    row += 1
    if quality.get("altman_capped"):
        ws.cell(row=row, column=1, value=(
            "Nota: il punteggio mostrato in pagina è limitato a 40 per zona di distress Altman — la "
            "formula sopra ricalcola il composito 'grezzo' pre-override."
        )).font = Font(name=FONT_NAME, italic=True, size=9, color=RED)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=(
        "I pesi di categoria dipendono dall'archetipo operativo (Dickinson + caratteristiche osservabili, "
        "src/lifecycle.py) e sono già cap-adjusted per il peso Piotroski; modificarli manualmente sopra "
        "rompe questa coerenza — utile solo per simulazioni 'what-if'."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    _autosize(ws, {"A": 44, "B": 22, "C": 20})


# ---------------------------------------------------------------------------
# Foglio 4 — Valuation (4 componenti)
# ---------------------------------------------------------------------------

def _write_valuation_sheet(ws, symbol: str, result: dict):
    ws.cell(row=1, column=1, value=f"Componenti Valuation — {symbol}").font = _title_font(14)
    ws.cell(row=2, column=1, value=(
        "Punteggio alto = economico. Nessun peer group: multipli assoluti calibrati per settore, storia "
        "propria (percentile su finestra storica), earnings yield vs risk-free, growth-adjusted."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    labels = {
        "sector_multiples": "Multipli assoluti (EV/EBITDA, EV/Sales, P/E) vs settore",
        "own_history": "Storia propria (percentile P/E)",
        "earnings_yield_vs_rf": "EV/EBIT earnings yield vs Treasury 10Y",
        "growth_adjusted": "Growth-adjusted (PEG o Rule of 40)",
    }
    header_row = 4
    for j, h in enumerate(["Componente", "Punteggio (0-100)"]):
        ws.cell(row=header_row, column=1 + j, value=h)
    _style_header_row(ws, header_row, 2)

    components = result.get("valuation", {}).get("components", {})
    row = header_row + 1
    for key, label in labels.items():
        ws.cell(row=row, column=1, value=label).font = _label_font()
        c = ws.cell(row=row, column=2, value=components.get(key))
        c.font = _input_font()
        c.number_format = "0.0"
        row += 1

    _autosize(ws, {"A": 50, "B": 16})


# ---------------------------------------------------------------------------
# Foglio 5 — Note Critiche
# ---------------------------------------------------------------------------

def _write_notes_sheet(ws, symbol: str, result: dict):
    ws.cell(row=1, column=1, value=f"Note Critiche — {symbol}").font = _title_font(14)
    ws.cell(row=2, column=1, value=(
        "Situazioni diagnosticabili in cui le metriche standard possono ingannare — mostrate solo "
        "quando il trigger machine-detectable scatta (src/critical_notes.py)."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    notes = result.get("critical_notes", [])
    if not notes:
        ws.cell(row=4, column=1, value="Nessuna nota critica scattata per questo titolo.").font = _label_font()
        _autosize(ws, {"A": 90})
        return

    header_row = 4
    for j, h in enumerate(["Codice", "Testo"]):
        ws.cell(row=header_row, column=1 + j, value=h)
    _style_header_row(ws, header_row, 2)

    row = header_row + 1
    for note in notes:
        ws.cell(row=row, column=1, value=note["code"]).font = Font(name=FONT_NAME, bold=True, color=RED)
        cell = ws.cell(row=row, column=2, value=note["text"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 45
        row += 1

    _autosize(ws, {"A": 12, "B": 100})


# ---------------------------------------------------------------------------
# Foglio 6 — Bilancio annuale (dati grezzi + margini calcolati)
# ---------------------------------------------------------------------------

def _period_labels(hist: dict) -> list[str]:
    for key, _ in _RAW_ROWS:
        s = hist.get(key)
        if s is not None and len(s) > 0:
            return [c.strftime("%Y-%m") if hasattr(c, "strftime") else str(c) for c in s.index]
    return []


def _write_raw_table(ws, start_row: int, hist: dict, title: str) -> int:
    labels = _period_labels(hist)
    ws.cell(row=start_row, column=1, value=title).font = _title_font(12)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Voce (dati storici Yahoo Finance)").font = _header_font()
    for j, lbl in enumerate(labels):
        ws.cell(row=header_row, column=2 + j, value=lbl)
    _style_header_row(ws, header_row, 1 + len(labels))

    row_map = {}
    r = header_row + 1
    for key, label in _RAW_ROWS:
        s = hist.get(key)
        ws.cell(row=r, column=1, value=label).font = _label_font()
        if s is not None:
            for j in range(len(labels)):
                val = float(s.iloc[j]) if j < len(s) else None
                cell = ws.cell(row=r, column=2 + j, value=val)
                cell.font = _input_font()
                cell.number_format = "$#,##0;($#,##0);-" if key != "eps" else "$#,##0.00;($#,##0.00);-"
        row_map[key] = r
        r += 1

    margin_defs = [
        ("gross_profit", "Margine lordo %"), ("operating_income", "Margine operativo %"),
        ("ebitda", "Margine EBITDA %"), ("net_income", "Margine netto %"),
    ]
    for num_key, label in margin_defs:
        if num_key not in row_map or "revenue" not in row_map:
            continue
        ws.cell(row=r, column=1, value=label).font = _label_font()
        num_row, rev_row = row_map[num_key], row_map["revenue"]
        for j in range(len(labels)):
            col = get_column_letter(2 + j)
            formula = f"=IFERROR({col}{num_row}/{col}{rev_row},\"n/d\")"
            cell = ws.cell(row=r, column=2 + j, value=formula)
            cell.font = _formula_font()
            cell.number_format = "0.0%"
        r += 1

    ws.cell(row=r, column=1, value=(
        "Fonte: Yahoo Finance (yfinance), prospetti contabili annuali. Le voci sopra sono dati storici "
        "(input, testo blu); i margini % sotto sono formule (testo nero)."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color=GRAY)

    widths = {"A": 30}
    for j in range(len(labels)):
        widths[get_column_letter(2 + j)] = 13
    _autosize(ws, widths)
    return r + 2


# ---------------------------------------------------------------------------
# Orchestrazione
# ---------------------------------------------------------------------------

def build_excel_report(symbol: str, info: dict, price, result: dict) -> bytes:
    """Costruisce il workbook completo dell'Analisi Fondamentale v2.0 e
    lo ritorna come bytes, pronto per `st.download_button`. `result` è
    l'output di `fundamental_score.build_fundamental_score()`."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Sintesi"
    _write_summary_sheet(ws_summary, symbol, info, price, result)

    ws_metrics = wb.create_sheet("Metriche core")
    _write_metrics_sheet(ws_metrics, symbol, result)

    ws_cat = wb.create_sheet("Quality - categorie e pesi")
    _write_categories_sheet(ws_cat, symbol, result)

    ws_val = wb.create_sheet("Valuation - componenti")
    _write_valuation_sheet(ws_val, symbol, result)

    ws_notes = wb.create_sheet("Note Critiche")
    _write_notes_sheet(ws_notes, symbol, result)

    ws_annual = wb.create_sheet("Bilancio annuale")
    hist = finmod.get_financial_history(symbol, freq="annual")
    _write_raw_table(ws_annual, 1, hist, f"Bilancio annuale — {symbol}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
